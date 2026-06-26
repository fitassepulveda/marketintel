#!/usr/bin/env python3
"""Scoring-insights report — reveals how the LLM is actually scoring.

Reads the articles the pipeline already scored, adds the six structured sub-scores
to any that lack them, then computes an INTERPRETABLE picture of the model's
behavior and emails it as an Excel workbook:

  1. What the model rewards   — each sub-dimension's influence on the LLM's 0-10
                                relevance score (correlation-based, robust at small
                                n; plus a regression view + R^2 when enough data).
  2. Dimension correlations   — 6x6 heatmap: which dimensions move together.
  3. Score distribution        — histogram of relevance scores (calibration check).
  4. By intelligence area      — average relevance + sub-scores per area.

IMPORTANT (for explaining upward): the LLM does NOT internally compute six numbers
and add them. The six sub-scores are a SECOND probing pass; this report explains the
model's single holistic score in interpretable human terms. It is an approximation
of a black box, not its literal internals.

  python scripts/scoring_insights.py --dry-run    # build workbook, don't email
  python scripts/scoring_insights.py --no-email    # alias for --dry-run
  python scripts/scoring_insights.py               # build + email to INSIGHTS_EMAIL_TO
"""
from __future__ import annotations
import argparse
import os
import smtplib
import sys
from datetime import datetime, timedelta, timezone
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import numpy as np
import yaml
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from src import ahp, config, store                  # noqa: E402
from src.llm_client import LLMClient                 # noqa: E402
from src.prioritize import subscores                 # noqa: E402


def preset_weights() -> dict:
    """The PRESET (prescriptive) weight per dimension — derived from our AHP pairwise
    judgments in config/ahp.yaml. This is 'what we said should matter', to compare
    against what the data shows actually drives the score."""
    try:
        pw = yaml.safe_load(open(config.CONFIG_DIR / "ahp.yaml")).get("pairwise", {})
        w = ahp.ahp_weights(ahp.matrix_from_pairwise(DIMS, pw))["weights"]
        return dict(zip(DIMS, w))
    except Exception:
        return {d: 1.0 / len(DIMS) for d in DIMS}

DIMS = subscores.DIMENSIONS
NAVY = "1F3864"
HEAD = Font(bold=True, color="FFFFFF", name="Arial")
HF = PatternFill("solid", start_color=NAVY)
BODY = Font(name="Arial")
DEFAULT_TO = "fxs1141@miami.edu"


# ----- data -----------------------------------------------------------------
def load_pool(con, window_days: int) -> list[dict]:
    import json
    since = (datetime.now(timezone.utc) - timedelta(days=window_days)).isoformat()
    rows = con.execute(
        "SELECT * FROM articles WHERE composite_score IS NOT NULL AND fetched >= ?",
        (since,)).fetchall()
    pool = []
    for r in rows:
        d = dict(r)
        if isinstance(d.get("subscores"), str):
            try:
                d["subscores"] = json.loads(d["subscores"])
            except Exception:
                d["subscores"] = None
        pool.append(d)
    return pool


def _complete(subs) -> bool:
    """True if a row's sub-scores cover every CURRENT dimension AND aren't all-zero.
    Adding new dimensions (missing keys) OR an all-zero set (the signature of a
    failed/throttled scoring batch) both count as incomplete, so they get re-scored
    rather than polluting the averages with zeros."""
    if not subs or (set(DIMS) - set(subs)):
        return False
    return any(float(v) > 0 for v in subs.values())


def backfill_subscores(con, client, cfg, pool: list[dict], max_items: int = 40):
    """Sub-score articles missing the current dimensions — but at most `max_items` per
    run (freshest first), so a one-time backlog (e.g. after adding dimensions) is spread
    over several runs instead of a single burst that trips Gemini's free-tier limits."""
    todo = [a for a in pool if not _complete(a.get("subscores"))]
    if not todo:
        return
    todo.sort(key=lambda a: (a.get("fetched") or ""), reverse=True)
    todo = todo[:max_items]
    models = cfg["settings"]["llm"]["models"][cfg["settings"]["llm"]["provider"]]
    res = subscores.score_batch(client, models["scoring"], cfg["settings"]["org"], todo)
    for art, ss in zip(todo, res):
        if not (ss and any(float(v) > 0 for v in ss.values())):
            continue  # failed/throttled batch came back all-zero — don't persist it
        subscores.save(con, art["id"], ss)
        art["subscores"] = ss
    con.commit()


# ----- analysis -------------------------------------------------------------
def mean_rewards(rows: list[dict]) -> dict:
    """Average sub-score (0-10) per dimension — 'how much the model rewards each
    dimension on average' — over the given rows."""
    rows = [r for r in rows if r.get("subscores")]
    if not rows:
        return {d: 0.0 for d in DIMS}
    return {d: float(np.mean([float(r["subscores"].get(d, 0)) for r in rows])) for d in DIMS}


def analyze(pool: list[dict]) -> dict:
    rows = [a for a in pool if a.get("subscores") and a.get("llm_score") is not None]
    n = len(rows)
    X = np.array([[float(a["subscores"].get(d, 0)) for d in DIMS] for a in rows]) if n else np.zeros((0, 6))
    y = np.array([float(a["llm_score"]) for a in rows]) if n else np.zeros(0)

    # influence via correlation of each dimension with the LLM relevance score
    corr_rel = np.zeros(len(DIMS))
    for j in range(len(DIMS)):
        if n >= 3 and X[:, j].std() > 0 and y.std() > 0:
            corr_rel[j] = np.corrcoef(X[:, j], y)[0, 1]
    corr_rel = np.nan_to_num(corr_rel)
    infl = np.abs(corr_rel)
    influence = infl / infl.sum() if infl.sum() > 0 else np.zeros(len(DIMS))

    # regression view (only meaningful with enough data)
    r2, coef, r2_ext = None, None, None
    if n >= 15:
        A = np.column_stack([X, np.ones(n)])
        beta, *_ = np.linalg.lstsq(A, y, rcond=None)
        coef = beta[:-1]
        pred = A @ beta
        ss_res = float(((y - pred) ** 2).sum())
        ss_tot = float(((y - y.mean()) ** 2).sum())
        r2 = 1 - ss_res / ss_tot if ss_tot > 0 else None

        # Extended: how much MORE variance is explained when we ALSO include the
        # intelligence area — a factor outside the six dimensions that influences the
        # score. The lift from r2 to r2_ext = how much "area" explains on its own.
        areas_u = sorted({row["area"] for row in rows})
        if len(areas_u) > 1 and ss_tot > 0:
            dummies = [np.array([1.0 if row["area"] == ar else 0.0 for row in rows]) for ar in areas_u[1:]]
            A2 = np.column_stack([X] + dummies + [np.ones(n)])
            b2, *_ = np.linalg.lstsq(A2, y, rcond=None)
            p2 = A2 @ b2
            r2_ext = 1 - float(((y - p2) ** 2).sum()) / ss_tot

    # 6x6 correlation matrix among dimensions
    cmat = np.corrcoef(X, rowvar=False) if n >= 3 else np.eye(len(DIMS))
    cmat = np.nan_to_num(cmat)

    # distribution of relevance scores (bins 0-1 .. 9-10)
    dist = [0] * 10
    for v in y:
        b = min(int(v), 9)
        dist[b] += 1

    # per-area
    areas = {}
    for a in rows:
        ar = a["area"]
        areas.setdefault(ar, []).append(a)
    by_area = []
    for ar, items in sorted(areas.items()):
        ys = np.array([float(i["llm_score"]) for i in items])
        means = {d: float(np.mean([float(i["subscores"].get(d, 0)) for i in items])) for d in DIMS}
        by_area.append({"area": ar, "n": len(items), "avg_relevance": float(ys.mean()), **means})

    return {"n": n, "corr_rel": corr_rel, "influence": influence, "coef": coef, "r2": r2,
            "r2_ext": r2_ext, "cmat": cmat, "dist": dist, "by_area": by_area,
            "avg_relevance": float(y.mean()) if n else 0.0}


# ----- workbook -------------------------------------------------------------
def _hdr(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEAD
        cell.fill = HF
        cell.alignment = Alignment(horizontal="center")


def build_workbook(a: dict, out: Path):
    wb = Workbook()

    # Sheet 1 — what the model rewards (avg score per dimension: last day vs all-time)
    means_last = a.get("means_last", {})
    means_all = a.get("means_all", {})
    infl = {d: float(i) for d, i in zip(DIMS, a["influence"])}
    preset = preset_weights()
    ws = wb.active
    ws.title = "What the Model Rewards"
    r2s = f"{a['r2']:.2f}" if a.get("r2") is not None else "n/a"
    r2e = f"{a['r2_ext']:.2f}" if a.get("r2_ext") is not None else "n/a"
    ws.append([f"Average reward per dimension (0-10 score the model gave). "
               f"Last day: {a.get('n_last', 0)} articles  |  All-time: {a.get('n_all', a['n'])} articles.   "
               f"Variance explained — dimensions: R²={r2s};  + intelligence area: R²={r2e}."])
    ws["A1"].font = Font(italic=True, name="Arial")
    ws.append(["Dimension", "Reward (last day)", "Reward (all-time)",
               "Influence on relevance", "Preset weight (AHP)"])
    _hdr(ws, 2, 5)
    ranked = sorted(DIMS, key=lambda d: means_all.get(d, 0), reverse=True)
    for d in ranked:
        ws.append([d, round(means_last.get(d, 0), 2), round(means_all.get(d, 0), 2),
                   round(infl.get(d, 0), 4), round(float(preset.get(d, 0)), 4)])
    for r in range(3, 3 + len(DIMS)):
        ws.cell(r, 1).font = BODY
        ws.cell(r, 2).font = BODY
        ws.cell(r, 3).font = BODY
        ws.cell(r, 4).number_format = "0.0%"; ws.cell(r, 4).font = BODY
        ws.cell(r, 5).number_format = "0.0%"; ws.cell(r, 5).font = BODY
    ws.column_dimensions["A"].width = 24
    for col in ("B", "C", "D", "E"):
        ws.column_dimensions[col].width = 18
    ch = BarChart()
    ch.type = "bar"
    ch.title = "What the model rewards — avg score per dimension (last day vs all-time)"
    ch.y_axis.title = "Average score (0-10)"
    ch.add_data(Reference(ws, min_col=2, max_col=3, min_row=2, max_row=2 + len(DIMS)), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=3, max_row=2 + len(DIMS)))
    ch.height, ch.width = 9, 18
    ws.add_chart(ch, "G2")

    # Sheet 2 — preset (what we SAID should matter) vs observed (what drives the score)
    wsP = wb.create_sheet("Preset vs Observed", 1)
    wsP.append(["Our PRESET priorities (AHP weights) vs what the data shows actually drives the score."])
    wsP["A1"].font = Font(italic=True, name="Arial")
    wsP.append(["Dimension", "Preset weight (AHP)", "Observed influence"])
    _hdr(wsP, 2, 3)
    for d in sorted(DIMS, key=lambda d: preset.get(d, 0), reverse=True):
        wsP.append([d, round(float(preset.get(d, 0)), 4), round(infl.get(d, 0), 4)])
    for r in range(3, 3 + len(DIMS)):
        wsP.cell(r, 1).font = BODY
        wsP.cell(r, 2).number_format = "0.0%"; wsP.cell(r, 2).font = BODY
        wsP.cell(r, 3).number_format = "0.0%"; wsP.cell(r, 3).font = BODY
    wsP.column_dimensions["A"].width = 24
    wsP.column_dimensions["B"].width = 20
    wsP.column_dimensions["C"].width = 20
    chP = BarChart(); chP.type = "bar"
    chP.title = "Preset priority (AHP) vs observed influence"
    chP.add_data(Reference(wsP, min_col=2, max_col=3, min_row=2, max_row=2 + len(DIMS)), titles_from_data=True)
    chP.set_categories(Reference(wsP, min_col=1, min_row=3, max_row=2 + len(DIMS)))
    chP.height, chP.width = 10, 18
    wsP.add_chart(chP, "E2")
    wsP.cell(3 + len(DIMS) + 1, 1,
             "Big gaps = where our stated priorities and the model's actual behavior diverge.").font = Font(italic=True, name="Arial")

    # Sheet 3 — dimension correlation heatmap
    ws2 = wb.create_sheet("Dimension Correlations")
    ws2.append([""] + DIMS)
    _hdr(ws2, 1, len(DIMS) + 1)
    for i, d in enumerate(DIMS):
        ws2.append([d] + [round(float(a["cmat"][i, j]), 2) for j in range(len(DIMS))])
        ws2.cell(2 + i, 1).font = HEAD
        ws2.cell(2 + i, 1).fill = HF
    last_col = chr(64 + 1 + len(DIMS))   # 'G' for 6 dims
    rng = f"B2:{last_col}{1 + len(DIMS)}"
    ws2.conditional_formatting.add(rng, ColorScaleRule(
        start_type="num", start_value=-1, start_color="F8696B",   # red
        mid_type="num", mid_value=0, mid_color="FFFFFF",          # white
        end_type="num", end_value=1, end_color="5A8AC6"))         # blue
    ws2.column_dimensions["A"].width = 24
    for c in range(2, len(DIMS) + 2):
        ws2.column_dimensions[chr(64 + c)].width = 12
    ws2.cell(len(DIMS) + 3, 1, "Read: +1 (blue) = move together, -1 (red) = move oppositely.").font = Font(italic=True, name="Arial")

    # Sheet 3 — score distribution
    ws3 = wb.create_sheet("Score Distribution")
    ws3.append(["Relevance score bin", "Article count"])
    _hdr(ws3, 1, 2)
    for i, c in enumerate(a["dist"]):
        ws3.append([f"{i}-{i+1}", c])
        ws3.cell(2 + i, 1).font = BODY
        ws3.cell(2 + i, 2).font = BODY
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 16
    ch3 = BarChart()
    ch3.title = "Distribution of LLM relevance scores"
    ch3.add_data(Reference(ws3, min_col=2, min_row=1, max_row=11), titles_from_data=True)
    ch3.set_categories(Reference(ws3, min_col=1, min_row=2, max_row=11))
    ch3.height, ch3.width = 9, 16
    ws3.add_chart(ch3, "D2")

    # Sheet 4 — by area
    ws4 = wb.create_sheet("By Area")
    ws4.append(["Intelligence area", "Articles", "Avg relevance"] + DIMS)
    _hdr(ws4, 1, 3 + len(DIMS))
    for row in a["by_area"]:
        ws4.append([row["area"], row["n"], round(row["avg_relevance"], 2)]
                   + [round(row[d], 1) for d in DIMS])
    ws4.column_dimensions["A"].width = 26
    for c in range(2, 4 + len(DIMS)):
        ws4.column_dimensions[chr(64 + c)].width = 13

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)


def summary_html(a: dict, date_h: str) -> str:
    ranked = sorted(zip(DIMS, a["influence"]), key=lambda x: -x[1])
    top = "".join(f"<li>{d.replace('_',' ')}: {inf*100:.0f}%</li>" for d, inf in ranked[:3])
    r2 = f"{a['r2']:.2f}" if a["r2"] is not None else "n/a (need >=15 articles)"
    return (
        f'<div style="font-family:Arial;max-width:640px">'
        f'<h2 style="color:#1F3864">Scoring Insights — {date_h}</h2>'
        f'<p>Based on <b>{a["n"]}</b> recently scored articles (avg relevance '
        f'{a["avg_relevance"]:.1f}/10). What the model rewarded most:</p><ol>{top}</ol>'
        f'<p>Regression fit R&sup2; = {r2}. Full breakdown — influence chart, 6&times;6 '
        f'correlation heatmap, score distribution, and per-area averages — is attached.</p>'
        f'<p style="color:#888;font-size:12px">These are interpretable approximations of the '
        f'model&rsquo;s single holistic 0&ndash;10 judgment, not its literal internals.</p></div>')


def email_workbook(path: Path, html: str, subject: str, to_addr: str):
    msg = MIMEMultipart()
    msg["Subject"] = subject
    msg["From"] = config.env("EMAIL_FROM")
    msg["To"] = to_addr
    msg.attach(MIMEText(html, "html"))
    with open(path, "rb") as f:
        part = MIMEApplication(f.read(), _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    part.add_header("Content-Disposition", "attachment", filename=path.name)
    msg.attach(part)
    with smtplib.SMTP(config.env("SMTP_HOST"), int(config.env("SMTP_PORT"))) as s:
        s.starttls()
        s.login(config.env("SMTP_USER"), config.env("SMTP_PASS"))
        s.sendmail(config.env("EMAIL_FROM"), [to_addr], msg.as_string())


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--no-email", action="store_true")
    ap.add_argument("--window-days", type=int, default=int(os.environ.get("INSIGHTS_WINDOW_DAYS", 7)))
    ap.add_argument("--max-subscore", type=int, default=int(os.environ.get("INSIGHTS_MAX_SUBSCORE", 40)),
                    help="cap how many articles get sub-scored per run (avoids free-tier rate limits)")
    args = ap.parse_args()

    cfg = config.load_all()
    con = store.connect()
    subscores.ensure_column(con)
    pool = load_pool(con, args.window_days)
    if len([a for a in pool if a.get("llm_score") is not None]) < 3:
        print(f"Only {len(pool)} scored articles in the last {args.window_days} days — "
              "need a few more days of runs before the insights are meaningful.")
        return

    client = LLMClient(cfg["settings"]["llm"]["provider"])
    backfill_subscores(con, client, cfg, pool, max_items=args.max_subscore)

    # All-time = every sub-scored article in the DB; last-day = those fetched in the
    # last 24h. The "What the Model Rewards" chart compares these two.
    # All-time uses only fully-scored rows (every current dimension present), so adding
    # new dimensions doesn't dilute their averages with zeros from older articles.
    all_pool = [r for r in subscores.load_scored(con)
                if r.get("llm_score") is not None and _complete(r.get("subscores"))]
    day_ago = (datetime.now(timezone.utc) - timedelta(days=1)).isoformat()
    last_pool = [r for r in all_pool if (r.get("fetched") or "") >= day_ago]

    a = analyze(all_pool)
    a["means_all"] = mean_rewards(all_pool)
    a["means_last"] = mean_rewards(last_pool)
    a["n_all"] = len(all_pool)
    a["n_last"] = len(last_pool)

    run_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    date_h = datetime.now().strftime("%A, %B %d, %Y")
    out = config.DATA_DIR / f"scoring_insights_{run_date}.xlsx"
    build_workbook(a, out)
    print(f"Wrote {out} ({a['n']} articles analyzed)")

    if args.dry_run or args.no_email:
        print("Not emailed (--dry-run/--no-email).")
        return
    to_addr = os.environ.get("INSIGHTS_EMAIL_TO", DEFAULT_TO)
    email_workbook(out, summary_html(a, date_h),
                   f'Scoring Insights — {date_h}', to_addr)
    print(f"Emailed insights workbook to {to_addr}")


if __name__ == "__main__":
    main()
