#!/usr/bin/env python3
"""Export AHP + eigen-analysis results to an Excel workbook with charts.

  python scripts/ahp_to_excel.py                  # -> data/ahp_results.xlsx
  python scripts/ahp_to_excel.py my_report.xlsx   # custom path

Sheets:
  - AHP Weights      prescriptive weights (principal eigenvector) + bar chart
  - Pairwise Matrix  the judgment matrix behind those weights
  - Consistency      lambda_max, CI, CR, consistent?  (audit trail)
  - Eigen-Analysis   (only if >=5 sub-scored articles exist) variance explained
                     scree, PC1 loadings, and mean sub-scores, each charted

Values are computed in Python (eigenvector decomposition isn't an Excel formula);
re-run this script after changing config/ahp.yaml or after more pipeline runs.
"""
import sys
from pathlib import Path

import numpy as np
import yaml
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from src import ahp, config, store  # noqa: E402
from src.prioritize import subscores  # noqa: E402

DIMS = subscores.DIMENSIONS
NAVY = "1F3864"
HEAD = Font(bold=True, color="FFFFFF", name="Arial")
HEAD_FILL = PatternFill("solid", start_color=NAVY)
BODY = Font(name="Arial")


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEAD
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(horizontal="center")


def sheet_weights(wb, weights):
    ws = wb.active
    ws.title = "AHP Weights"
    ws.append(["Dimension", "Weight"])
    _style_header(ws, 1, 2)
    ranked = sorted(zip(DIMS, weights), key=lambda x: -x[1])
    for dim, w in ranked:
        ws.append([dim, round(float(w), 4)])
    for r in range(2, 2 + len(DIMS)):
        ws.cell(row=r, column=1).font = BODY
        ws.cell(row=r, column=2).font = BODY
        ws.cell(row=r, column=2).number_format = "0.0%"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 12

    chart = BarChart()
    chart.type = "bar"
    chart.title = "AHP Priority Weights (principal eigenvector)"
    chart.y_axis.title = "Weight"
    data = Reference(ws, min_col=2, min_row=1, max_row=1 + len(DIMS))
    cats = Reference(ws, min_col=1, min_row=2, max_row=1 + len(DIMS))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height, chart.width = 9, 18
    ws.add_chart(chart, "D2")


def sheet_matrix(wb, matrix):
    ws = wb.create_sheet("Pairwise Matrix")
    ws.append([""] + DIMS)
    _style_header(ws, 1, len(DIMS) + 1)
    for i, dim in enumerate(DIMS):
        row = [dim] + [round(float(matrix[i, j]), 3) for j in range(len(DIMS))]
        ws.append(row)
        ws.cell(row=2 + i, column=1).font = HEAD
        ws.cell(row=2 + i, column=1).fill = HEAD_FILL
    ws.column_dimensions["A"].width = 24
    for col in range(2, len(DIMS) + 2):
        ws.column_dimensions[chr(64 + col)].width = 13


def sheet_consistency(wb, matrix):
    res = ahp.ahp_weights(matrix)
    ws = wb.create_sheet("Consistency")
    rows = [
        ("Metric", "Value"),
        ("Matrix size (n)", len(DIMS)),
        ("lambda_max", round(res["lambda_max"], 4)),
        ("Consistency Index (CI)", round(res["consistency_index"], 4)),
        ("Consistency Ratio (CR)", round(res["consistency_ratio"], 4)),
        ("Acceptable? (CR < 0.10)", "YES" if res["consistent"] else "NO — revise judgments"),
    ]
    for r in rows:
        ws.append(list(r))
    _style_header(ws, 1, 2)
    for r in range(2, 2 + len(rows) - 1):
        ws.cell(row=r, column=1).font = BODY
        ws.cell(row=r, column=2).font = BODY
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22


def sheet_eigen(wb):
    con = store.connect()
    subscores.ensure_column(con)
    rows = subscores.load_scored(con)
    if len(rows) < 5:
        ws = wb.create_sheet("Eigen-Analysis")
        ws["A1"] = ("Need >= 5 sub-scored articles for the data-driven eigen-analysis. "
                    "Run run_personalized.py a few times, then re-run this export.")
        ws["A1"].font = BODY
        ws.column_dimensions["A"].width = 90
        return

    data = np.array([[r["subscores"].get(d, 0.0) for d in DIMS] for r in rows])
    res = ahp.eigen_analysis(data)

    ws = wb.create_sheet("Eigen-Analysis")
    ws.append([f"Based on {len(rows)} sub-scored articles"])
    ws["A1"].font = Font(italic=True, name="Arial")

    ws.append(["Principal Component", "Variance Explained"])
    _style_header(ws, 2, 2)
    for i, ve in enumerate(res["variance_explained"]):
        ws.append([f"PC{i+1}", round(float(ve), 4)])
        ws.cell(row=3 + i, column=2).number_format = "0.0%"
    scree_start = 3
    scree_end = 2 + len(res["variance_explained"])

    head2 = scree_end + 2
    ws.cell(row=head2, column=1, value="Dimension")
    ws.cell(row=head2, column=2, value="PC1 Loading")
    ws.cell(row=head2, column=3, value="Mean Sub-score")
    _style_header(ws, head2, 3)
    means = data.mean(axis=0)
    ranked = sorted(zip(DIMS, res["loadings"], means), key=lambda x: -x[1])
    for k, (dim, load, mean) in enumerate(ranked):
        r = head2 + 1 + k
        ws.cell(row=r, column=1, value=dim).font = BODY
        ws.cell(row=r, column=2, value=round(float(load), 4)).number_format = "0.0%"
        ws.cell(row=r, column=3, value=round(float(mean), 2))
    load_start, load_end = head2 + 1, head2 + len(DIMS)

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16

    scree = BarChart()
    scree.title = "Variance Explained (scree)"
    scree.add_data(Reference(ws, min_col=2, min_row=2, max_row=scree_end), titles_from_data=True)
    scree.set_categories(Reference(ws, min_col=1, min_row=scree_start, max_row=scree_end))
    scree.height, scree.width = 8, 14
    ws.add_chart(scree, "E2")

    loadc = BarChart()
    loadc.type = "bar"
    loadc.title = "PC1 Loadings by dimension"
    loadc.add_data(Reference(ws, min_col=2, min_row=head2, max_row=load_end), titles_from_data=True)
    loadc.set_categories(Reference(ws, min_col=1, min_row=load_start, max_row=load_end))
    loadc.height, loadc.width = 9, 14
    ws.add_chart(loadc, "E20")


def main():
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else config.DATA_DIR / "ahp_results.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)

    pairwise = yaml.safe_load(open(config.CONFIG_DIR / "ahp.yaml")).get("pairwise", {})
    matrix = ahp.matrix_from_pairwise(DIMS, pairwise)
    weights = ahp.ahp_weights(matrix)["weights"]

    wb = Workbook()
    sheet_weights(wb, weights)
    sheet_matrix(wb, matrix)
    sheet_consistency(wb, matrix)
    sheet_eigen(wb)
    wb.save(out)
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
