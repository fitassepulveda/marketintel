#!/usr/bin/env python3
"""Competitor new-development / new-construction report.

Pulls every South Florida competitor FOOTPRINT move out of data/intel.db —
construction, M&A, partnerships, capital gifts, new service lines — dedupes the
repeat coverage down to one row per PROJECT, and writes an Excel tracker.

    # build the workbook
    python scripts/competitor_development_report.py

    # build + email it to yourself (needs SMTP_* + EMAIL_FROM in .env)
    python scripts/competitor_development_report.py --email wef28@miami.edu

    # widen/narrow the window (default: everything in the archive)
    python scripts/competitor_development_report.py --since 2026-06-01

COVERAGE CAVEAT: intel.db only holds what the pipeline has collected since
2026-06-15 (competitor-newsroom items reach back to ~April). This is NOT a
12-month history — the Coverage tab states the real window.

Re-running is safe and idempotent: the CURATION table below pins the rows that
have been human-reviewed, and anything NEW that the filters catch lands in the
"Unreviewed Candidates" tab for you to promote.
"""
from __future__ import annotations

import argparse
import re
import sqlite3
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

DB = ROOT / "data" / "intel.db"

# --------------------------------------------------------------------------
# Competitor set — the eight in config/settings.yaml, plus Broward Health and a
# market bucket. Patterns match hospital brands, not just system names, because
# coverage often names the facility ("Bethesda", "Joe DiMaggio") and never the parent.
# --------------------------------------------------------------------------
COMPETITORS = {
    "Baptist Health": r"baptist health|baptist hospital|bethesda hospital|boca raton regional|"
                      r"miami cancer institute|miami cardiac|marcus neuroscience|lynn cancer|"
                      r"wertheim cancer institute|homestead hospital|west kendall baptist",
    "Jackson Health": r"jackson health|jackson memorial|jackson north|jackson south|jackson west|"
                      r"holtz children|jackson heart institute|public health trust",
    "Cleveland Clinic Florida": r"cleveland clinic",
    "Memorial Healthcare": r"memorial healthcare|memorial regional|memorial hospital (west|miramar|pembroke)|"
                           r"joe dimaggio|south broward hospital district",
    "HCA Florida": r"hca florida|hca healthcare|kendall regional|aventura hospital|mercy hospital|"
                   r"jfk (medical|north)|northwest medical center|westside regional|plantation general|"
                   r"palms west",
    "Holy Cross Health": r"holy cross",
    "Mount Sinai Medical Center": r"mount sinai",
    "Nicklaus Children's": r"nicklaus children",
    "Broward Health": r"broward health",
    "Other market": r"florida coast medical center|steward health|orlando health|tampa general|"
                    r"adventhealth|tenet health|florida international university|\bfiu\b|"
                    r"miami dade college|larkin|palmetto general",
}

# Construction / physical-footprint signals ONLY. Bricks and mortar: new buildings,
# groundbreakings, openings, renovations, land buys. Partnerships, payer deals and
# philanthropy not tied to a specific building are deliberately NOT here.
DEV_SIGNAL = (
    r"construct|break(s|ing)? ground|broke ground|groundbreak|ground-break|ribbon cutting|"
    r"topping out|cement pour|new hospital|new campus|new tower|new building|new facility|"
    r"medical office building|freestanding|free-standing|micro-hospital|"
    r"new emergency|new er\b|new ed\b|emergency room|emergency department|"
    r"now open|opens |opening|will open|unveil|"
    r"expansion|expanding|expand|renovat|remodel|refurbish|build(s|ing)?\b|"
    r"square[- ]foot|square feet|sq\.? ?ft|\bbeds\b|"
    r"property sale|land (buy|purchase|acquisition)|acquires? (the )?(property|site|land|campus)|"
    r"breaks? ground|relocat|new location|new site|new clinic|new center\b|new institute"
)

# --------------------------------------------------------------------------
# Geography. South Florida = Miami-Dade, Broward, Palm Beach, Monroe.
# The 35-mile band is UHealth's PPS-exempt cancer-center service area, measured
# from the Miami Health District; moves outside it are real but weigh less.
# --------------------------------------------------------------------------
CITY_GEO = {
    # city -> (county, approx miles from UM/Miller School Health District)
    "Miami": ("Miami-Dade", 0), "Miami Health District": ("Miami-Dade", 0),
    "North Miami": ("Miami-Dade", 10), "Doral": ("Miami-Dade", 9),
    "Aventura": ("Miami-Dade", 17), "Kendall": ("Miami-Dade", 13),
    "Homestead": ("Miami-Dade", 30), "Coral Gables": ("Miami-Dade", 5),
    "Miami Beach": ("Miami-Dade", 6), "West Miami-Dade": ("Miami-Dade", 10),
    "Miramar": ("Broward", 22), "Pembroke Pines": ("Broward", 20),
    "Hollywood": ("Broward", 18), "Sunrise": ("Broward", 28),
    "Fort Lauderdale": ("Broward", 25), "Plantation": ("Broward", 27),
    "Weston": ("Broward", 30), "Tamarac": ("Broward", 33),
    "Boca Raton": ("Palm Beach", 44), "West Palm Beach": ("Palm Beach", 65),
    "Boynton Beach": ("Palm Beach", 52), "Palm Beach": ("Palm Beach", 66),
}
BAND_MILES = 35


def geo(city: str):
    if city == "Regionwide":
        return "Multi-county", "—", "Regionwide"
    if not city:
        return "", "—", "Unconfirmed"
    county, miles = CITY_GEO.get(city, ("", None))
    if miles is None:
        return county, "—", "Unconfirmed"
    return county, f"~{miles} mi", ("In band" if miles <= BAND_MILES else "Outside band")


# --------------------------------------------------------------------------
# CURATION — human-reviewed rows. Each entry pins one PROJECT and lists the
# article ids in intel.db that cover it, so repeat syndication collapses to one row.
# To add a project: append here with its article ids. To kill a false positive:
# add its ids to EXCLUDE with a reason.
# --------------------------------------------------------------------------
CURATED = [
    # --- Baptist Health -------------------------------------------------
    dict(comp="Baptist Health", ids=[160, 112],
         project="Helen and Jacob Shaham Academic Medical Center (FIU Modesto Maidique campus)",
         dtype="New construction — academic medical center", city="West Miami-Dade",
         invest="$158M", size="163,000 sq ft / 7 stories", status="Broke ground May 2026; opens 2028",
         note="Direct hit on UHealth's academic-medicine moat. Baptist buys an academic affiliation "
              "and a teaching campus in west Miami-Dade — an expansion zone — funded entirely by a "
              "state appropriation. Outpatient multi-specialty, same-day surgery, imaging, infusion."),
    dict(comp="Baptist Health", ids=[192, 2162, 60], scope="other",
         project="Herbert Wertheim $100M gift — Wertheim Cancer Institute Transformational Impact Fund",
         dtype="Capital gift — cancer program", city="Miami",
         invest="$100M", size="", status="Announced May 2026",
         note="Largest gift in Baptist's 166-year history and explicitly aimed at NCI designation — "
              "a direct challenge to Sylvester's NCI-designated status. Also deepens the FIU "
              "Wertheim College of Medicine tie and funds a new Center for Cancer Prevention."),
    dict(comp="Baptist Health", ids=[59, 1036, 80], scope="other",
         project="Amazon One Medical strategic partnership",
         dtype="Strategic partnership — primary care access", city="Regionwide",
         invest="", size="One Medical's 20-region network", status="Effective Aug 1, 2026",
         note="Baptist outsources front-door primary care to Amazon and keeps the specialty "
              "referrals. A hybrid virtual/in-person funnel feeding Baptist specialists is a "
              "referral-capture play UHealth has no equivalent to."),
    dict(comp="Baptist Health", ids=[193], scope="other",
         project="START Center for Cancer Research — new South Florida trial site",
         dtype="Research partnership — new site", city="Miami",
         invest="", size="", status="Announced May 2026",
         note="Brings early-phase oncology trials to Baptist patients who would otherwise travel "
              "out of region — erodes a classic academic-center differentiator for Sylvester."),
    dict(comp="Baptist Health", ids=[364],
         project="Slattery $1M gift — Homestead Hospital capacity expansion",
         dtype="Capital gift — facility expansion", city="Homestead",
         invest="$1M", size="", status="Announced May 2026",
         note="Small dollars, but signals continued investment in south Miami-Dade capacity."),

    # --- Jackson Health -------------------------------------------------
    dict(comp="Jackson Health", ids=[63, 64, 195, 196, 922, 974, 1902],
         project="Jackson Memorial Hospital Emergency Department expansion (Phase One)",
         dtype="New construction — emergency department", city="Miami Health District",
         invest="$400M ($100M philanthropic)", size="178,000+ sq ft; 60 exam rooms, → 121 adult + 30 pediatric",
         status="Phase One opened Apr 23, 2026; completion announced Jul 2026",
         note="The single biggest capacity move in the market, and it is across the street from "
              "UHealth. Doubles Jackson's ED footprint, adds behavioral-health bays and full "
              "advanced imaging. Reshapes emergency referral flow in the Health District."),
    dict(comp="Jackson Health", ids=[66],
         project="Jackson Heart Institute cardiology hub, North Miami",
         dtype="New ambulatory site — service line", city="North Miami",
         invest="", size="", status="Open",
         note="Geographic expansion of a flagship service line straight into a UHealth expansion "
              "zone. General + interventional cardiology, EP and heart failure under one roof."),
    dict(comp="Jackson Health", ids=[65, 197],
         project="Health Science Collegiate Academy with Miami Dade College",
         dtype="Workforce partnership — new facility", city="Miami Health District",
         invest="", size="", status="Binding agreement; opens Aug 2027",
         note="A charter high school in the Health District that pipelines nursing and allied "
              "health talent to Jackson first. Workforce is a named UHealth strategic vector."),

    # --- Memorial Healthcare -------------------------------------------
    dict(comp="Memorial Healthcare", ids=[763],
         project="Red Road ER, Miramar",
         dtype="New freestanding ER", city="Miramar",
         invest="", size="", status="Opened Jun 30, 2026",
         note="Freestanding ER pushing hospital-level emergency care into south Broward — part of "
              "a visible Memorial land-grab on access points."),
    dict(comp="Memorial Healthcare", ids=[1982, 1979],
         project="Douglas Road ER, Pembroke Pines",
         dtype="New freestanding ER (converted 24/7 care center)", city="Pembroke Pines",
         invest="", size="", status="Opened Jul 21, 2026",
         note="Second freestanding ER in three weeks. Memorial is converting existing urgent-care "
              "sites into full ERs — fast, cheap capacity adds that capture acuity and admissions."),
    dict(comp="Memorial Healthcare", ids=[2088, 2079], scope="other",
         project="Aetna multi-year contract renewal",
         dtype="Payer agreement", city="Regionwide",
         invest="", size="", status="Renewed Jul 22, 2026",
         note="Locks in in-network access across Memorial's hospitals and specialists. Removes a "
              "network-disruption opening UHealth could otherwise have played for."),
    dict(comp="Memorial Healthcare", ids=[1061], scope="other",
         project="Joe DiMaggio first remote living-donor kidney transplant",
         dtype="Service-line capability — first in program", city="Hollywood",
         invest="", size="", status="Jul 6, 2026",
         note="Pediatric transplant capability is a quaternary service line where academic centers "
              "usually hold the edge. Worth tracking as a capability creep signal."),

    # --- Broward Health -------------------------------------------------
    dict(comp="Broward Health", ids=[277],
         project="Medical office building + parking garage, Broward Health Medical Center",
         dtype="New construction — MOB", city="Fort Lauderdale",
         invest="", size="", status="Cement pour milestone Jun 23, 2026",
         note="Outpatient capacity next to the flagship — the standard ambulatory-shift build."),
    dict(comp="Broward Health", ids=[926],
         project="Broward Health Sunrise ER (second freestanding ER)",
         dtype="New freestanding ER", city="Sunrise",
         invest="", size="", status="Ribbon cutting Jul 1, 2026",
         note="First ER in the City of Sunrise — greenfield access point, not a replacement. "
              "Broward and Memorial are both racing on freestanding ERs."),
    dict(comp="Broward Health", ids=[459],
         project="School Board property acquisition",
         dtype="Land acquisition", city="Fort Lauderdale",
         invest="", size="", status="Approved Jun 25, 2026",
         note="Land banking. Worth watching for what gets announced on the parcel."),
    dict(comp="Broward Health", ids=[1540],
         project="Renovated Labor & Delivery suite, Broward Health Medical Center",
         dtype="Renovation — service line", city="Fort Lauderdale",
         invest="", size="", status="Opened Jul 14, 2026",
         note="Maternal health repositioning at the flagship."),

    # --- Holy Cross -----------------------------------------------------
    dict(comp="Holy Cross Health", ids=[841],
         project="New $57M healthcare facility",
         dtype="New construction", city="Fort Lauderdale",
         invest="$57M", size="", status="Broke ground Jul 1, 2026",
         note="Trinity Health capital landing in Broward. Details thin in the archive — worth a "
              "targeted follow-up to pin the site and service mix."),

    # --- HCA Florida ----------------------------------------------------
    dict(comp="HCA Florida", ids=[1358],
         project="Palms West Hospital freestanding ER, West Palm Beach",
         dtype="New freestanding ER", city="West Palm Beach",
         invest="", size="", status="Broke ground Jul 10, 2026",
         note="OUTSIDE the 35-mile band — weighted a notch lower. Still the clearest read on HCA's "
              "freestanding-ER playbook, which is the format most likely to appear in Miami-Dade next."),

    # --- Nicklaus Children's --------------------------------------------
    dict(comp="Nicklaus Children's", ids=[381], scope="other",
         project="Region's first pediatric total hip replacement (mixed-reality surgical guidance)",
         dtype="Service-line capability — first in region", city="Miami",
         invest="", size="", status="Jun 23, 2026",
         note="Orthopedic, Sports Medicine & Spine Institute claiming a regional first with "
              "mixed-reality surgical guidance. Pediatric quaternary orthopedics is contested "
              "ground; Nicklaus is using technology firsts to hold referral share."),

    # --- Cleveland Clinic Florida --------------------------------------
    dict(comp="Cleveland Clinic Florida", ids=[244],
         project="Medical office project advancing (Florida city, unspecified in archive)",
         dtype="New construction — medical office", city="",
         invest="", size="", status="Moving forward Jun 20, 2026",
         note="LOW CONFIDENCE: headline-only Google News item, city not captured. Needs manual "
              "follow-up before it is briefed — the underlying article names the municipality."),
    dict(comp="Cleveland Clinic Florida", ids=[1815], scope="other",
         project="Palm Beach gala — $22M raised",
         dtype="Philanthropy — capital", city="Palm Beach",
         invest="$22M", size="", status="Jul 20, 2026",
         note="Outside the 35-mile band. Confirms Cleveland Clinic's donor base in Palm Beach, "
              "which historically funds its Weston/Palm Beach county expansion."),
]

# False positives and out-of-region items, kept visible for auditability.
EXCLUDE = {
    2090: ("Cleveland Clinic Florida", "Haslam $25M gift", "Enterprise/Ohio gift, not the Florida region"),
    2076: ("Cleveland Clinic Florida", "Haslam $25M gift (dup)", "Enterprise/Ohio gift, not the Florida region"),
    463: ("Cleveland Clinic Florida", "First US robotic lung transplant", "Main campus (Ohio) procedure"),
    458: ("Cleveland Clinic Florida", "First US robotic lung transplant (dup)", "Main campus (Ohio) procedure"),
    693: ("HCA Florida", "Brandon medical offices ribbon cutting", "Tampa market, not South Florida"),
    690: ("HCA Florida", "Raulerson leadless pacemaker", "Okeechobee, not South Florida"),
    1905: ("HCA Florida", "Lehigh stroke certification", "Lehigh Acres (SW FL), not South Florida"),
    278: ("Other market", "Florida Coast Medical Center STEMI", "Treasure Coast, outside South Florida"),
    1990: ("Other market", "AdventHealth/Intermountain JV", "Denver market"),
    1975: ("Other market", "Intermountain/AdventHealth JV (dup)", "Denver market"),
    2057: ("Other market", "Intermountain/AdventHealth JV (dup)", "Denver market"),
    236: ("Baptist Health", "Cayman Islands JCI re-accreditation", "Accreditation, not a footprint move; non-SF"),
    611: ("Baptist Health", "Baptist Health Richmond CMO", "Baptist Health Kentucky — name collision"),
    369: ("Mount Sinai Medical Center", "AI portfolio oversight", "Mount Sinai New York — name collision"),
    444: ("Mount Sinai Medical Center", "DOJ records", "Mount Sinai New York — name collision"),
    105: ("Mount Sinai Medical Center", "Wisp PrEP partnership", "Mount Sinai New York (NY market)"),
    34: ("Mount Sinai Medical Center", "Wisp PrEP partnership (dup)", "Mount Sinai New York (NY market)"),
    1192: ("HCA Florida", "Mission Health BSN requirement", "North Carolina"),
    231: ("Baptist Health", "Habitat for Humanity 100 homes", "Community investment — outside report scope"),
    280: ("Cleveland Clinic Florida", "Food security community investment", "Community investment — outside report scope"),
}


def load_articles(conn, since: str | None):
    q = ("select id, title, summary, source, url, published, area, llm_score, composite_score "
         "from articles")
    rows = {r["id"]: dict(r) for r in conn.execute(q)}
    if since:
        rows = {i: r for i, r in rows.items()
                if not (r["published"] or "") or (r["published"] or "")[:10] >= since}
    return rows


def auto_candidates(rows):
    """Every article that looks like a competitor footprint move."""
    out = []
    for r in rows.values():
        blob = f"{r['title']} {r['summary'] or ''}".lower()
        if not re.search(DEV_SIGNAL, blob):
            continue
        hits = [k for k, p in COMPETITORS.items() if re.search(p, blob)]
        if not hits:
            continue
        out.append((r, hits))
    return out


def build_rows(rows, scope="build"):
    """scope='build' -> physical footprint only; scope='other' -> the non-construction moves."""
    curated_ids, out = set(), []
    for c in CURATED:
        arts = [rows[i] for i in c["ids"] if i in rows]
        curated_ids.update(c["ids"])
        if c.get("scope", "build") != scope:
            continue
        if not arts:
            continue
        dates = sorted({(a["published"] or "")[:10] for a in arts if (a["published"] or "").strip()})
        scores = [a["llm_score"] for a in arts if a["llm_score"] is not None]
        county, dist, band = geo(c["city"])
        out.append({
            "Competitor": c["comp"],
            "First reported": dates[0] if dates else "undated (competitor newsroom)",
            "Latest coverage": dates[-1] if dates else "",
            "Project": c["project"],
            "Development type": c["dtype"],
            "City": c["city"],
            "County": county,
            "Distance from UHealth": dist,
            "35-mi band": band,
            "Investment": c["invest"],
            "Size / capacity": c["size"],
            "Status": c["status"],
            "Why it matters to UHealth": c["note"],
            "Peak LLM score": max(scores) if scores else "",
            "Articles": len(arts),
            "Source": arts[0]["source"],
            "URL": arts[0]["url"],
            "Article IDs": ", ".join(str(i) for i in c["ids"]),
        })
    order = {k: i for i, k in enumerate(COMPETITORS)}
    out.sort(key=lambda r: (order.get(r["Competitor"], 99), r["First reported"]))
    return out, curated_ids


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--since", help="only articles published on/after YYYY-MM-DD")
    ap.add_argument("--out", default=None, help="output .xlsx path")
    ap.add_argument("--email", nargs="*", help="email the workbook to these addresses")
    args = ap.parse_args()

    if not DB.exists():
        sys.exit(f"No database at {DB}")

    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    rows = load_articles(conn, args.since)
    dev_rows, curated_ids = build_rows(rows, scope="build")
    other_rows, _ = build_rows(rows, scope="other")
    cands = auto_candidates(rows)
    unreviewed = [(r, h) for r, h in cands
                  if r["id"] not in curated_ids and r["id"] not in EXCLUDE]

    pubs = sorted({(r["published"] or "")[:10] for r in rows.values() if (r["published"] or "").strip()})
    coverage = (pubs[0], pubs[-1]) if pubs else ("", "")

    out_path = Path(args.out) if args.out else (
        ROOT.parent / f"competitor_developments_{date.today():%Y-%m-%d}.xlsx")
    write_workbook(out_path, dev_rows, unreviewed, rows, coverage, len(rows), other_rows)
    print(f"Wrote {out_path}  ({len(dev_rows)} construction/footprint projects, "
          f"{len(other_rows)} non-construction moves, {len(unreviewed)} unreviewed candidates)")

    if args.email is not None:
        send(out_path, args.email or ["wef28@miami.edu"], dev_rows, coverage)


def write_workbook(path, dev_rows, unreviewed, rows, coverage, n_articles, other_rows=()):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    BRAND, INK, MUTED = "006888", "313A45", "6B7480"
    hdr_fill = PatternFill("solid", fgColor=BRAND)
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()

    # ---- Developments -----------------------------------------------
    ws = wb.active
    ws.title = "Developments"
    cols = list(dev_rows[0].keys()) if dev_rows else []
    ws.append(cols)
    for r in dev_rows:
        ws.append([r[c] for c in cols])
    widths = {"Competitor": 22, "First reported": 14, "Latest coverage": 14, "Project": 52,
              "Development type": 34, "City": 18, "County": 12, "Distance from UHealth": 12,
              "35-mi band": 13, "Investment": 20, "Size / capacity": 34, "Status": 30,
              "Why it matters to UHealth": 72, "Peak LLM score": 9, "Articles": 8,
              "Source": 26, "URL": 42, "Article IDs": 20}
    for i, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 16)
    for cell in ws[1]:
        cell.fill, cell.font, cell.alignment = hdr_fill, hdr_font, Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    band_col = cols.index("35-mi band") + 1 if "35-mi band" in cols else None
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(cols)):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(size=10, color=INK)
        if band_col and row[band_col - 1].value == "Outside band":
            row[band_col - 1].font = Font(size=10, color="B45309", bold=True)
    # Shade alternating competitor blocks so the grouping reads at a glance.
    tint = PatternFill("solid", fgColor="F2F7F9")
    shade, prev = False, None
    for excel_row, r in enumerate(dev_rows, start=2):
        if r["Competitor"] != prev:
            shade, prev = not shade, r["Competitor"]
        if shade:
            for cell in ws[excel_row]:
                cell.fill = tint
        ws.cell(row=excel_row, column=1).font = Font(size=10, bold=True, color=BRAND)
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"

    # ---- One tab per competitor -------------------------------------
    # EVERY competitor gets a tab, including those where nothing qualified — the
    # "Everything else pulled" block below shows what the pipeline saw and why each
    # item did or didn't become a development row. No silent gaps.
    sub_cols = [c for c in cols if c != "Competitor"]
    sub_widths = {c: widths.get(c, 16) for c in sub_cols}
    curated_ids = {i for c in CURATED if c.get("scope", "build") == "build" for i in c["ids"]}
    other_ids = {i for c in CURATED if c.get("scope", "build") == "other" for i in c["ids"]}
    dev_signal_ids = {r["id"] for r, _ in auto_candidates(rows)}

    for comp, pattern in COMPETITORS.items():
        comp_rows = [r for r in dev_rows if r["Competitor"] == comp]
        pulled = [a for a in rows.values()
                  if re.search(pattern, f"{a['title']} {a['summary'] or ''}".lower())]
        tab = re.sub(r"[\\/*?:\[\]]", "", comp)[:31]
        wsc = wb.create_sheet(tab)

        in_band = sum(1 for r in comp_rows if r["35-mi band"] == "In band")
        wsc.append([comp])
        wsc.append([f"{len(comp_rows)} qualifying development{'s' if len(comp_rows) != 1 else ''} "
                    f"· {in_band} inside the 35-mile band · {len(pulled)} articles pulled in total"])
        wsc["A1"].font = Font(bold=True, size=14, color=BRAND)
        wsc["A2"].font = Font(size=10, color=MUTED)
        wsc.append([])

        hdr_row = wsc.max_row + 1
        if comp_rows:
            wsc.append(sub_cols)
            for r in comp_rows:
                wsc.append([r[c] for c in sub_cols])
            for i, c in enumerate(sub_cols, 1):
                wsc.column_dimensions[get_column_letter(i)].width = sub_widths[c]
            for cell in wsc[hdr_row]:
                cell.fill, cell.font = hdr_fill, hdr_font
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            wsc.row_dimensions[hdr_row].height = 30
            sband = sub_cols.index("35-mi band") + 1
            for row in wsc.iter_rows(min_row=hdr_row + 1, max_row=wsc.max_row, max_col=len(sub_cols)):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    cell.font = Font(size=10, color=INK)
                if row[sband - 1].value == "Outside band":
                    row[sband - 1].font = Font(size=10, color="B45309", bold=True)
            wsc.freeze_panes = f"A{hdr_row + 1}"
        else:
            if comp == "Other market":
                why = ("Catch-all bucket for non-core players and academic partners. The FIU and "
                       "Miami Dade College projects it matches are attributed to Baptist and "
                       "Jackson respectively, so they are not double-counted here. Everything "
                       "else is out of region (Tampa, Treasure Coast, Denver).")
            elif pulled and all(a["id"] in EXCLUDE for a in pulled):
                why = ("Every item pulled was a name collision with a same-named system outside "
                       "South Florida — see dispositions below.")
            else:
                why = "Nothing in the archive met the footprint criteria for this competitor."
            wsc.append(["No qualifying development or footprint move captured in this window."])
            wsc.append([why])
            wsc[f"A{hdr_row}"].font = Font(bold=True, size=11, color="B45309")
            wsc[f"A{hdr_row + 1}"].font = Font(size=10, color=INK)
            for i, c in enumerate(sub_cols, 1):
                wsc.column_dimensions[get_column_letter(i)].width = sub_widths[c]

        # --- Everything else pulled on this competitor, with disposition ---
        # openpyxl's max_row ignores appended blank rows, so step down explicitly.
        sec = wsc.max_row + 3
        wsc[f"A{sec}"] = "Everything else pulled on this competitor"
        wsc[f"A{sec}"].font = Font(bold=True, size=11, color=BRAND)
        sub_hdr = sec + 1
        for j, h in enumerate(["Article ID", "Date", "Title", "Source", "LLM score",
                               "Disposition", "URL"], 1):
            wsc.cell(row=sub_hdr, column=j, value=h)
        for cell in wsc[sub_hdr][:7]:
            cell.fill, cell.font = PatternFill("solid", fgColor=MUTED), hdr_font

        others = [a for a in pulled if a["id"] not in curated_ids]
        rowi = sub_hdr
        for a in sorted(others, key=lambda x: (x["published"] or "")):
            if a["id"] in EXCLUDE:
                disp = f"Excluded — {EXCLUDE[a['id']][2]}"
            elif a["id"] in other_ids:
                disp = "Strategic move, no construction — see Non-Construction Moves tab"
            elif a["id"] in dev_signal_ids:
                disp = "Candidate — needs review"
            else:
                disp = "Not a build (leadership, clinical, community, or coverage noise)"
            rowi += 1
            for j, v in enumerate([a["id"], (a["published"] or "")[:10], a["title"], a["source"],
                                   a["llm_score"], disp, a["url"]], 1):
                wsc.cell(row=rowi, column=j, value=v)
        if not others:
            wsc.cell(row=sub_hdr + 1, column=3,
                     value="Nothing else pulled — all coverage is captured above.")
        for row in wsc.iter_rows(min_row=sub_hdr + 1, max_row=wsc.max_row, max_col=7):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.font = Font(size=9, color=INK)
            if str(row[5].value or "").startswith("Excluded"):
                row[5].font = Font(size=9, color="B45309")
            elif str(row[5].value or "").startswith("Candidate"):
                row[5].font = Font(size=9, color=BRAND, bold=True)

    # ---- Summary by competitor --------------------------------------
    ws2 = wb.create_sheet("Summary")
    agg = defaultdict(lambda: {"n": 0, "types": set(), "band": 0, "cap": 0.0})
    for r in dev_rows:
        a = agg[r["Competitor"]]
        a["n"] += 1
        a["types"].add(re.sub(r"\s*\(.*?\)", "", r["Development type"].split(" — ")[0]).strip())
        if r["35-mi band"] == "In band":
            a["band"] += 1
        m = re.search(r"\$(\d+(?:\.\d+)?)\s*([MB])", r["Investment"] or "", re.I)
        if m:
            a["cap"] += float(m.group(1)) * (1000 if m.group(2).upper() == "B" else 1)
    ws2.append(["Competitor", "Projects captured", "Inside 35-mi band", "Disclosed capital ($M)",
                "Articles pulled", "Move types"])
    for comp, pattern in COMPETITORS.items():
        a = agg.get(comp, {"n": 0, "types": set(), "band": 0, "cap": 0.0})
        pulled = sum(1 for x in rows.values()
                     if re.search(pattern, f"{x['title']} {x['summary'] or ''}".lower()))
        ws2.append([comp, a["n"], a["band"], round(a["cap"], 1) or "", pulled,
                    ", ".join(sorted(a["types"])) or "— no qualifying footprint move —"])
    ws2.append(["TOTAL", sum(a["n"] for a in agg.values()),
                sum(a["band"] for a in agg.values()),
                round(sum(a["cap"] for a in agg.values()), 1), "", ""])
    for i, w in enumerate([28, 18, 18, 22, 16, 58], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    for cell in ws2[1]:
        cell.fill, cell.font = hdr_fill, hdr_font
    for cell in ws2[ws2.max_row]:
        cell.font = Font(bold=True, color=INK)

    # ---- Non-construction moves (parked, not counted) ----------------
    wsn = wb.create_sheet("Non-Construction Moves")
    wsn["A1"] = "Strategic moves with no physical build"
    wsn["A1"].font = Font(bold=True, size=14, color=BRAND)
    wsn["A2"] = ("Out of scope for this report, kept for reference. Partnerships, payer contracts, "
                 "philanthropy not tied to a specific building, and service-line firsts.")
    wsn["A2"].font = Font(size=10, color=MUTED)
    nc_cols = [c for c in ("Competitor", "First reported", "Project", "Development type", "City",
                           "Investment", "Status", "Why it matters to UHealth", "Source", "URL")]
    for j, h in enumerate(nc_cols, 1):
        wsn.cell(row=4, column=j, value=h)
    for k, r in enumerate(other_rows, start=5):
        for j, c in enumerate(nc_cols, 1):
            wsn.cell(row=k, column=j, value=r[c])
    for i, c in enumerate(nc_cols, 1):
        wsn.column_dimensions[get_column_letter(i)].width = widths.get(c, 16)
    for cell in wsn[4][:len(nc_cols)]:
        cell.fill, cell.font = hdr_fill, hdr_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    wsn.row_dimensions[4].height = 30
    for row in wsn.iter_rows(min_row=5, max_row=wsn.max_row, max_col=len(nc_cols)):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(size=10, color=INK)
    wsn.freeze_panes = "A5"

    # ---- Unreviewed candidates --------------------------------------
    ws3 = wb.create_sheet("Unreviewed Candidates")
    ws3.append(["Article ID", "Date", "Competitor match", "Title", "Source", "LLM score", "URL", "Keep? (y/n)"])
    for r, hits in sorted(unreviewed, key=lambda x: (x[0]["published"] or "")):
        ws3.append([r["id"], (r["published"] or "")[:10], "; ".join(hits), r["title"],
                    r["source"], r["llm_score"], r["url"], ""])
    for i, w in enumerate([10, 12, 26, 70, 26, 10, 40, 12], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    for cell in ws3[1]:
        cell.fill, cell.font = hdr_fill, hdr_font
    ws3.freeze_panes = "A2"

    # ---- Excluded ----------------------------------------------------
    ws4 = wb.create_sheet("Excluded")
    ws4.append(["Article ID", "Competitor", "Item", "Reason excluded"])
    for aid, (comp, item, reason) in sorted(EXCLUDE.items()):
        ws4.append([aid, comp, item, reason])
    for i, w in enumerate([10, 26, 46, 60], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w
    for cell in ws4[1]:
        cell.fill, cell.font = hdr_fill, hdr_font

    # ---- Coverage & method -------------------------------------------
    ws5 = wb.create_sheet("Coverage & Method")
    notes = [
        ("Report", "Competitor new development & construction — South Florida"),
        ("Generated", f"{date.today():%Y-%m-%d}"),
        ("Source", "marketintel/data/intel.db — the market intel pipeline archive only"),
        ("", ""),
        ("COVERAGE CAVEAT", "This is NOT a 12-month history."),
        ("Archive window", f"{coverage[0]} to {coverage[1]} (dated articles)"),
        ("Pipeline collecting since", "2026-06-15; competitor-newsroom items backfill to ~April 2026"),
        ("Implication", "Anything before roughly April 2026 was never captured and is absent here. "
                        "A true 12-month view requires web research beyond the archive."),
        ("", ""),
        ("Articles scanned", n_articles),
        ("Scope", "CONSTRUCTION AND PHYSICAL FOOTPRINT ONLY: new buildings, groundbreakings, "
                  "freestanding ERs, medical office buildings, hospital/ED expansions, renovations, "
                  "new clinical sites, and land acquisitions for future build."),
        ("Out of scope", "Partnerships, payer contracts, philanthropy not tied to a specific "
                         "building, service-line firsts, leadership changes, community sponsorships, "
                         "accreditations, research findings. Strategic non-build moves are parked on "
                         "the Non-Construction Moves tab rather than deleted."),
        ("Competitors", ", ".join(COMPETITORS)),
        ("", ""),
        ("35-mile band", "UHealth's PPS-exempt cancer-center service area, measured from the Miami "
                         "Health District. Moves outside it are flagged and weigh less."),
        ("Location check", "Each row was confirmed to be the South Florida institution. Name "
                           "collisions (Baptist Kentucky, Mount Sinai New York, Cleveland Clinic Ohio) "
                           "are listed on the Excluded tab."),
        ("", ""),
        ("Dedupe", "One row per PROJECT. Repeat syndication is collapsed; the Article IDs column "
                   "lists every underlying record."),
        ("Re-running", "python scripts/competitor_development_report.py — new matches land on the "
                       "Unreviewed Candidates tab; promote them into the CURATED list in the script."),
    ]
    for k, v in notes:
        ws5.append([k, v])
    ws5.column_dimensions["A"].width = 28
    ws5.column_dimensions["B"].width = 100
    for row in ws5.iter_rows(min_row=1, max_row=ws5.max_row):
        row[0].font = Font(bold=True, color=BRAND, size=10)
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
        row[1].font = Font(size=10, color=INK)
    ws5["B5"].font = Font(size=10, bold=True, color="B45309")

    # Tab order: All Developments, Summary, one tab per competitor, then appendices.
    comp_tabs = [re.sub(r"[\\/*?:\[\]]", "", c)[:31] for c in COMPETITORS]
    desired = (["Developments", "Summary"]
               + [t for t in comp_tabs if t in wb.sheetnames]
               + ["Non-Construction Moves", "Unreviewed Candidates", "Excluded", "Coverage & Method"])
    wb._sheets = [wb[t] for t in desired if t in wb.sheetnames]

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def send(path, recipients, dev_rows, coverage):
    """Email the workbook using the same SMTP settings run_briefing.py uses."""
    import smtplib
    from email.mime.application import MIMEApplication
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from src import config

    need = ("SMTP_HOST", "SMTP_PORT", "SMTP_USER", "SMTP_PASS", "EMAIL_FROM")
    if not all(config.env(k, required=False) for k in need):
        print("SMTP not configured (set SMTP_* + EMAIL_FROM in .env) — workbook saved, not sent.")
        return

    by_comp = defaultdict(int)
    for r in dev_rows:
        by_comp[r["Competitor"]] += 1
    lines = "".join(f"<li><b>{c}</b> — {n} project{'s' if n != 1 else ''}</li>"
                    for c, n in by_comp.items())
    html = f"""
    <div style="font-family:Arial,Helvetica,sans-serif;color:#313A45;font-size:14px;line-height:1.5">
      <h2 style="color:#006888;margin:0 0 4px">Competitor New Development &amp; Construction — South Florida</h2>
      <p style="color:#6B7480;margin:0 0 16px">{len(dev_rows)} projects captured &middot; archive window {coverage[0]} to {coverage[1]}</p>
      <ul>{lines}</ul>
      <p style="background:#FFF7ED;border-left:3px solid #B45309;padding:10px 12px;margin:16px 0">
        <b>Coverage caveat:</b> built from the market intel archive only, which has been collecting
        since 2026-06-15 (competitor newsrooms backfill to ~April 2026). This is not a 12-month history.
      </p>
      <p>Full tracker attached — Developments, Summary, Unreviewed Candidates, Excluded, and Coverage &amp; Method tabs.</p>
    </div>"""

    msg = MIMEMultipart()
    msg["Subject"] = f"Competitor New Development Report — South Florida ({date.today():%b %d, %Y})"
    msg["From"] = config.env("EMAIL_FROM")
    msg["To"] = ", ".join(recipients)
    msg.attach(MIMEText(html, "html"))
    with open(path, "rb") as f:
        att = MIMEApplication(f.read(), _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    att.add_header("Content-Disposition", "attachment", filename=path.name)
    msg.attach(att)

    with smtplib.SMTP(config.env("SMTP_HOST"), int(config.env("SMTP_PORT"))) as s:
        s.starttls()
        s.login(config.env("SMTP_USER"), config.env("SMTP_PASS"))
        s.sendmail(config.env("EMAIL_FROM"), recipients, msg.as_string())
    print(f"Emailed to {', '.join(recipients)}")


if __name__ == "__main__":
    main()
